import re
import datetime
from pprint import pprint

import webest2 as w
from tinydb import TinyDB, Query

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


URL_SEARCH = 'https://www.mobile.de/es/veh%C3%ADculos/buscar.html?sb=doc&od=down&vc=Car&ao=PICTURES&c=OffRoad&c=Limousine&cn=DE&con=USED&dam=0&ft=ELECTRICITY&ml=%3A10000&p=%3A60000&re=400&s=Car'

# Create DB object
db_fn = f"cars-{datetime.now().strftime('%Y-%m-%d_%H%M')}.json"
db_search = TinyDB(db_fn)


def __clean_car_url(url):
    """
    Eg: /es/veh%C3%ADculos/detalles.html?id=398179510&amp;sb=doc&amp;od=down&amp;vc=Car&amp;ao=PICTURES&amp;c=OffRoad&amp;c=Limousine&amp;cn=DE&amp;con=USED&amp;dam=0&amp;ft=ELECTRICITY&amp;ml=%3A10000&amp;p=%3A60000&amp;re=400&amp;s=Car&amp;searchId=aaa037d9-136d-f2ef-608b-9db657f55ac0&amp;ref=srp&amp;refId=aaa037d9-136d-f2ef-608b-9db657f55ac0
                                                     ^
    """
    return url.split('&')[0]


def __dump_html(o):
    return o.get_attribute('outerHTML')


def _parse_price_eur(s: str) -> int | None:
    """
    Parse price strings like '49.547 €' or '49.547\\xa0€' → 49547 (int).
    
    Rules:
    - Value must be a non-negative integer with '.' as thousands separator only.
    - Must end with '€' (euro symbol), optionally preceded by whitespace (including \\xa0).
    - No decimals, no commas, no extra text.
    
    Valid:
        '0 €'                → 0
        '1 €'                → 1
        '1.000 €'            → 1000
        '49.547\\xa0€'       → 49547
        '1.000.000 €'        → 1000000
        '  500 €  '          → 500
    
    Invalid → None:
        '49.547'             (no €)
        '49,547 €'           (comma)
        '49.547.50 €'        (decimal-like — 2-digit group at end)
        '49.54 €'            (only 2 digits after dot — invalid grouping)
        '€49.547'            (€ before number)
        '49.547 USD'         (wrong currency)
        '49.547 € extra'     (extra text)
        '49.547. €'          (trailing dot)
    """
    # Pattern:
    # ^\s*                                - leading whitespace
    # (\d{1,3}(?:\.\d{3})*)              - integer with valid dot-thousands groups
    # [ \xa0]+                           - one or more spaces or non-breaking spaces
    # €                                  - euro symbol
    # \s*$                               - optional trailing whitespace, end
    pattern = r"^\s*(\d{1,3}(?:\.\d{3})*)[ \xa0]+€\s*$"
    match = re.fullmatch(pattern, s)
    
    if not match:
        return None
    
    num_str = match.group(1)
    
    try:
        clean_num = num_str.replace('.', '')
        value = int(clean_num)
        return value
    except (ValueError, OverflowError):
        return None


def _parse_km(s: str) -> int | None:
    """
    Parse strings like '8.000 km' → 8000 (int), where '.' is a thousands separator.
    Only non-negative integers allowed. No decimals.
    
    Valid: '8 km', '1.000 km', '12.345 km', '1.000.000 km'
    Invalid (→ None): '8. km', '.500 km', '8.00.0 km', '8,000 km', floats, extra text, etc.
    """
    # Strict full-match pattern:
    # - Optional leading/trailing whitespace
    # - Group 1: digits optionally grouped by dots (e.g., 1 or more dot-separated digit groups)
    #   Must start and end with digits; no leading/trailing dots
    # - Then 1+ whitespace, then 'km' (case-insensitive)
    pattern = r"^\s*(\d{1,3}(?:\.\d{3})*)\s+km\s*$"
    match = re.fullmatch(pattern, s, re.IGNORECASE)
    
    if not match:
        return None
    
    num_str = match.group(1)
    
    try:
        # Remove dots and parse as int
        clean_num = num_str.replace('.', '')
        value = int(clean_num)
        return value
    except (ValueError, OverflowError):
        return None


def _parse_kw(s: str) -> int | None:
    """
    Parse power strings like '350 kW (476 cv)' → 350 (int).
    
    Rules:
    - kW value must be a non-negative integer with optional '.' as thousands separator.
    - Must contain 'kW' (case-insensitive), preceded by the number and whitespace.
    - Optional trailing parenthetical info (e.g., '(476 cv)') is allowed and ignored.
    - No floats, no commas, no extra text outside the pattern.
    
    Valid:
        '350 kW'
        '350 kW (476 cv)'
        '1.200 kW'
        '1.200 kW (1.632 cv)'
        '  75 kW  '
    
    Invalid → None:
        '350kw' (no space)
        '350.5 kW' (decimal)
        '350 kW extra'
        '350 hp'
        'three hundred kW'
        '35,0 kW'
    """
    # Pattern explanation:
    # ^\s*                                - leading whitespace
    # (\d{1,3}(?:\.\d{3})*)              - integer with optional dot-thousands groups (e.g., 350, 1.200)
    # \s+                                 - at least one space
    # kW                                  - literal 'kW', case-insensitive
    # (?:\s*\(.*?\))?                    - optional: space + parenthetical (e.g., (476 cv))
    # \s*$                                - optional trailing whitespace, end
    pattern = r"^\s*(\d{1,3}(?:\.\d{3})*)\s+kW(?:\s*\(.*?\))?\s*$"
    match = re.fullmatch(pattern, s, re.IGNORECASE)
    
    if not match:
        return None
    
    num_str = match.group(1)
    
    try:
        clean_num = num_str.replace('.', '')
        value = int(clean_num)
        return value
    except (ValueError, OverflowError):
        return None


def _parse_minutes(s: str) -> int | None:
    """
    Parse time strings like '18 Min.' → 18 (int).
    
    Rules:
    - Value must be a non-negative integer; '.' used only as thousands separator.
    - Unit must be 'min' (case-insensitive), optionally followed by a dot.
    - No extra text before or after (except optional whitespace).
    
    Valid:
        '18 Min.'
        '18 min'
        '18 MIN'
        '1.200 min.'
        '0 Min'
        '  5 min.  '
    
    Invalid → None:
        '18.5 min'       (float)
        '18mins'         (no space)
        '18 min extra'   (extra text)
        'eighteen min'   (non-numeric)
        '18,000 min'     (comma not allowed)
        '18'             (unit missing)
    """
    # Pattern:
    # ^\s*                                - leading whitespace
    # (\d{1,3}(?:\.\d{3})*)              - integer with optional dot-thousands groups
    # \s+                                 - at least one space
    # min\.?                             - 'min' + optional trailing dot
    # \s*$                               - optional trailing whitespace, end
    pattern = r"^\s*(\d{1,3}(?:\.\d{3})*)\s+min\.?\s*$"
    match = re.fullmatch(pattern, s, re.IGNORECASE)
    
    if not match:
        return None
    
    num_str = match.group(1)
    
    try:
        clean_num = num_str.replace('.', '')
        value = int(clean_num)
        return value
    except (ValueError, OverflowError):
        return None


def _parse_int(s: str) -> int | None:
    """
    Parse owner count strings like '1' → 1 (int).
    
    Rules:
    - Must be a non-negative integer (0, 1, 2, ...)
    - No thousands separators (e.g., '1.000' is invalid — owners are small integers)
    - No units, no extra text, no signs, no decimals
    - Leading/trailing whitespace is stripped and allowed
    
    Valid:
        '0'      → 0
        '1'      → 1
        '2'      → 2
        '  3  '  → 3
    
    Invalid → None:
        '1 owner'    (extra text)
        '1.0'        (float)
        '1.'         (trailing dot)
        '.5'         (invalid)
        'two'        (non-numeric)
        ''           (empty)
        '-'          (invalid)
        '1,000'      (comma)
        '1 2'        (multiple numbers)
    """
    # Match optional whitespace, then digits only, then optional whitespace — nothing else.
    # Use fullmatch for strictness.
    match = re.fullmatch(r"\s*(\d+)\s*", s)
    
    if not match:
        return None
    
    num_str = match.group(1)
    
    # Guard against leading zeros (e.g., '01') unless it's just '0'
    # Optional: allow '01' → 1? Most systems disallow leading zeros in counts.
    if len(num_str) > 1 and num_str.startswith('0'):
        return None  # e.g., '01', '00' → invalid; only '0' is OK for zero.
    
    try:
        value = int(num_str)
        return value
    except (ValueError, OverflowError):
        return None


def perform_search():
    w.load(URL_SEARCH)

    # Read pagination
    pagination = w.get_obj('ul[class*="pagination_Pagination__"] li:nth-last-child(2)')
    pages_num = int(pagination.text)
    print(f"📚 {pages_num} result pages")

    for pag_num in range(1, pages_num + 1):
        if not (cont := perform_search_in_pag_num(pag_num)):
            break


def perform_search_in_pag_num(pag_num):
    Result = Query()

    url_search_pag = URL_SEARCH + f'&pageNumber={pag_num}'
    w.load(url_search_pag)
    print(f"🤓 Reading {url_search_pag}")

    # Page
    n_adds = n_skips = 0
    for o in w.get_objs('a[class*="BaseListing_containerLink"]'):
        car_url = __clean_car_url(o.get_attribute("href"))

        if db_search.get(Result.URL == car_url) is None:
            db_search.insert({'URL': car_url, 'needs_details': True})
            print(f"➕ Adds {car_url}")
            n_adds += 1
        else:
            print(f"➖ Skips {car_url}")
            n_skips += 1

    return n_adds >= 1


def fetch_details():
    Result = Query()

    cars_to_update = list(db_search.search(Result.needs_details == True))

    for car_entry in cars_to_update:
        car_url = car_entry['URL']

        print(f"📰 Fetching {car_url}")
        w.load(car_url)

        o = w.get_obj('h2[class*="typography_headline"]')
        car_title = o.get_attribute('textContent')

        o = w.get_obj('div[class*="MainCtaBox_subTitle"]')
        car_subtitle = o.get_attribute('textContent') if o else ''

        o = w.get_obj('div[class*="MainPriceArea_mainPrice__"]')
        car_price = o.get_attribute('textContent')

        o = w.get_obj('div[class*="priceRatingBadge_PriceRatingBadge--label_"]')
        car_price_fairness = o.get_attribute('textContent')

        car_data = {
            'title': car_title,
            'subtitle': car_subtitle,
            'price': _parse_price_eur(car_price),
            'price fairness': car_price_fairness
        }

        # Fetch all info pairs
        info_pairs_raw = {}
        for o in w.get_objs('div[class*="KeyFeatures_content__"]'):
            key_value = [s.strip() for s in o.text.split("\n")]
            info_pairs_raw[key_value[0]] = key_value[1]

        # Parse all the info pairs
        info_pairs = {}
        if (val := info_pairs_raw.get('Kilometraje')):
            info_pairs['Kilometraje'] = _parse_km(val)

        if (val := info_pairs_raw.get('Potencia')):
            info_pairs['Potencia (Kw)'] = _parse_kw(val)
            info_pairs['Potentia detalle'] = val

        if (val := info_pairs_raw.get('Tiempo de carga rápida')):
            info_pairs['Tiempo de carga rápida (mins)'] = _parse_minutes(val)

        if (val := info_pairs_raw.get('Autonomía (WLTP)')):
            info_pairs['Autonomía WLTP (Km)'] = _parse_km(val)

        if (val := info_pairs_raw.get('Propietarios anteriores')):
            info_pairs['Propietarios anteriores'] = _parse_int(val)
        
        # Update Database (the entry being iterated on)
        car_data = car_data | info_pairs
        car_data['needs_details'] = False
        db_search.update(car_data, doc_ids=[car_entry.doc_id])


def list_updated():
    Result = Query()

    for car_entry in db_search.search(Result.needs_details == False):
        pprint(car_entry)


def cli_update(skip_search=False, skip_details=False):
    with w.init_context():
        if not skip_search:
            perform_search()
        if not skip_details:
            fetch_details()


def cli_ls():
    list_updated()


def cli_sheet(fp_sheet="coches.xlsx"):
    Result = Query()

    car_list = []
    for car_entry in db_search.search(Result.needs_details == False):
        del(car_entry['needs_details'])
        car_list.append(car_entry)

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Coches Electricos"

    # Define column order (customize this to your preference)
    column_order = [
        'title', 'subtitle', 'price', 'price fairness', 
        'Kilometraje', 'Autonomía WLTP (Km)', 'Potencia (Kw)', 
        'Potentia detalle', 'Propietarios anteriores', 
        'URL'
    ]        

    # Write headers
    for col_idx, header in enumerate(column_order, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)  # Bold headers
        
    # Write data rows
    for row_idx, car in enumerate(car_list, start=2):
        for col_idx, field in enumerate(column_order, 1):
            value = car.get(field, "")
            
            # Special handling for URL field
            if field == 'URL' and isinstance(value, str):
                clean_url = value.strip()
                cell = ws.cell(row=row_idx, column=col_idx, value=clean_url)
                if clean_url:
                    cell.hyperlink = clean_url
                    cell.font = Font(color="0000FF", underline="single")  # Blue underlined text
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Set column width with padding (max 50 characters)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Save the file
    wb.save(fp_sheet)
    print(f"💾 Spreadsheet created successfully: {fp_sheet}")


if __name__ == "__main__":
    import fire
    fire.Fire({
        'update': cli_update,
        'ls': cli_ls,
        'sheet': cli_sheet,
    })